#!/usr/bin/env python3
"""Build ENGG433/956 Week 4 tutorial Excel workbooks."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

# Colours (Wiley-like)
NAVY = "1F4E79"
BLUE = "2E75B6"
LTBLUE = "D6EAF8"
PALE = "DEEBF7"
YELLOW = "FFF2CC"
GREEN = "C6EFCE"
ORANGE = "FCE4D6"
GREY = "F2F2F2"
WHITE = "FFFFFF"
BLACK = "000000"
DKGREEN = "375623"
RED = "C00000"

thin = Border(
    left=Side(style="thin", color="BDD7EE"),
    right=Side(style="thin", color="BDD7EE"),
    top=Side(style="thin", color="BDD7EE"),
    bottom=Side(style="thin", color="BDD7EE"),
)
med = Border(
    left=Side(style="thin", color="1F4E79"),
    right=Side(style="thin", color="1F4E79"),
    top=Side(style="thin", color="1F4E79"),
    bottom=Side(style="thin", color="1F4E79"),
)
bottom_double = Border(
    left=Side(style="thin", color="BDD7EE"),
    right=Side(style="thin", color="BDD7EE"),
    top=Side(style="thin", color="1F4E79"),
    bottom=Side(style="double", color="1F4E79"),
)
bottom_single = Border(
    left=Side(style="thin", color="BDD7EE"),
    right=Side(style="thin", color="BDD7EE"),
    top=Side(style="thin", color="1F4E79"),
    bottom=Side(style="medium", color="1F4E79"),
)

fill_navy = PatternFill("solid", fgColor=NAVY)
fill_blue = PatternFill("solid", fgColor=BLUE)
fill_lt = PatternFill("solid", fgColor=LTBLUE)
fill_pale = PatternFill("solid", fgColor=PALE)
fill_yellow = PatternFill("solid", fgColor=YELLOW)
fill_green = PatternFill("solid", fgColor=GREEN)
fill_orange = PatternFill("solid", fgColor=ORANGE)
fill_grey = PatternFill("solid", fgColor=GREY)
fill_white = PatternFill("solid", fgColor=WHITE)

font_title = Font(name="Calibri", size=18, bold=True, color=NAVY)
font_subtitle = Font(name="Calibri", size=12, italic=True, color=BLUE)
font_h = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_h2 = Font(name="Calibri", size=12, bold=True, color=NAVY)
font_label = Font(name="Calibri", size=11, bold=True, color=NAVY)
font_body = Font(name="Calibri", size=11, color=BLACK)
font_small = Font(name="Calibri", size=10, italic=True, color="666666")
font_formula = Font(name="Calibri", size=10, italic=True, color=BLUE)
font_ans = Font(name="Calibri", size=11, bold=True, color=DKGREEN)
font_money = Font(name="Calibri", size=11)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center")


def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def header_bar(ws, row, start, end, text, fill=fill_navy, font=font_h, height=22):
    ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
    cell = ws.cell(row, start, text)
    cell.font = font
    cell.fill = fill
    cell.alignment = center
    cell.border = med
    for c in range(start + 1, end + 1):
        ws.cell(row, c).fill = fill
        ws.cell(row, c).border = med
    ws.row_dimensions[row].height = height
    return cell


def money(cell, fmt='$#,##0.0;($#,##0.0);"—"'):
    cell.number_format = fmt
    cell.font = font_money
    cell.alignment = Alignment(horizontal="right", vertical="center")


def pct(cell):
    cell.number_format = "0.0%"
    cell.alignment = Alignment(horizontal="center", vertical="center")


def paint(cell, fill=None, font=None, align=None, border=None):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = border


# ---------------------------------------------------------------------------
# Workbook 1: Cash flow template (E17.2 + E17.12)
# ---------------------------------------------------------------------------
def build_cashflow():
    wb = Workbook()

    # ---- Cover ----
    ws = wb.active
    ws.title = "Cover"
    set_col_widths(ws, {"A": 4, "B": 28, "C": 55, "D": 28, "E": 18})
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B2:E2")
    ws["B2"] = "ENGG433 / ENGG956  ·  Financial Management for Engineers"
    ws["B2"].font = font_subtitle
    ws.merge_cells("B3:E3")
    ws["B3"] = "Week 4 Tutorial  ·  Statement of Cash Flows"
    ws["B3"].font = font_title
    ws.merge_cells("B4:E4")
    ws["B4"] = "Cash-flow statement template  ·  E17.2  and  E17.12"
    ws["B4"].font = Font(name="Calibri", size=13, color=BLUE)

    notes = [
        ("Sheet", "What to do"),
        ("E17.2 Classification", "Classify each Hailey Corp. item into one of four SCF categories (indirect method)."),
        ("E17.12 Direct Method", "Compute cash payments to suppliers and cash payments for operating expenses."),
        ("How to use", "Yellow cells are inputs already filled from the tutorial PDF. Green cells are formula answers — do not overwrite."),
        ("Indirect vs Direct", "Indirect starts from net income and adjusts. Direct reconstructs cash paid/received line by line. Both give the same operating cash flow."),
        ("Four SCF sections", "Operating  ·  Investing  ·  Financing  ·  Significant non-cash investing & financing (footnote / schedule)."),
    ]
    header_bar(ws, 6, 2, 3, "Workbook map")
    r = 7
    for a, b in notes:
        ws.cell(r, 2, a).font = font_label
        ws.cell(r, 2).fill = fill_lt
        ws.cell(r, 2).alignment = left
        ws.cell(r, 2).border = thin
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws.cell(r, 3, b).font = font_body
        ws.cell(r, 3).alignment = left
        ws.cell(r, 3).border = thin
        ws.cell(r, 4).border = thin
        ws.row_dimensions[r].height = 36
        r += 1

    ws.merge_cells("B14:D14")
    ws["B14"] = "Source: Weygandt, Kimmel, Kieso — E17.2 (LO1) and E17.12 (LO4). Figures in E17.12 are in $ millions."
    ws["B14"].font = font_small

    # ---- E17.2 ----
    ws = wb.create_sheet("E17.2 Classification")
    set_col_widths(ws, {"A": 5, "B": 42, "C": 48, "D": 18, "E": 55, "F": 22})
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"
    ws.merge_cells("B2:E2")
    ws["B2"] = "E17.2 (LO 1)  ·  Classify transactions by type of activity"
    ws["B2"].font = font_title
    ws.merge_cells("B3:E3")
    ws["B3"] = "Hailey Corp.  ·  Statement of Cash Flows (indirect method)"
    ws["B3"].font = font_subtitle
    ws.merge_cells("B4:E4")
    ws["B4"] = (
        "Indicate where each item is presented. Assume all items involve cash unless stated otherwise. "
        "Operating items are those listed among the adjustments to net income."
    )
    ws["B4"].font = font_small
    ws.row_dimensions[4].height = 32

    headers = ["Item", "Transaction", "Classification (answer)", "Cash effect", "Why / how it appears"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(7, i + 1, h)
        paint(cell, fill_navy, font_h, center, med)
    ws.row_dimensions[7].height = 24

    # answers
    rows = [
        (
            "a",
            "Exchange of land for patent",
            "Significant noncash investing and financing activity",
            "No cash",
            "Land given up, patent received — both investing assets, no cash. Disclose in a separate schedule (or notes).",
        ),
        (
            "b",
            "Sale of building at book value",
            "Investing activity",
            "Inflow",
            "Cash proceeds from disposing of a long-lived asset. Sold at book value → no gain/loss to adjust in operating.",
        ),
        (
            "c",
            "Payment of dividends",
            "Financing activity",
            "Outflow",
            "Cash returned to owners. (Declared but unpaid dividends are not a cash flow.)",
        ),
        (
            "d",
            "Depreciation of plant assets",
            "Operating activity",
            "Non-cash add-back",
            "Expense reduced NI but used no cash. Under the indirect method, ADD depreciation back to net income.",
        ),
        (
            "e",
            "Conversion of bonds into common stock",
            "Significant noncash investing and financing activity",
            "No cash",
            "Liability extinguished by issuing equity. Financing on both sides; disclose, do not put in the body of the SCF.",
        ),
        (
            "f",
            "Issuance of capital stock",
            "Financing activity",
            "Inflow",
            "Cash received from owners for shares issued.",
        ),
        (
            "g",
            "Amortization of patent",
            "Operating activity",
            "Non-cash add-back",
            "Same logic as depreciation: non-cash expense. ADD amortization back to net income.",
        ),
        (
            "h",
            "Issuance of bonds for land",
            "Significant noncash investing and financing activity",
            "No cash",
            "Investing (land acquired) + financing (bonds issued) in one non-cash deal. Disclose in the schedule.",
        ),
        (
            "i",
            "Purchase of land",
            "Investing activity",
            "Outflow",
            "Cash paid to acquire a long-lived productive asset.",
        ),
        (
            "j",
            "Loss on disposal of plant assets",
            "Operating activity",
            "Non-cash add-back",
            "Loss reduced NI but is not an operating cash outflow. ADD the loss back. (The cash proceeds, if any, go to Investing.)",
        ),
        (
            "k",
            "Retirement of bonds",
            "Financing activity",
            "Outflow",
            "Cash used to repay long-term debt principal. (Any gain/loss on extinguishment is an operating adjustment.)",
        ),
    ]

    classes = [
        "Operating activity",
        "Investing activity",
        "Financing activity",
        "Significant noncash investing and financing activity",
    ]
    dv = DataValidation(type="list", formula1='"' + ",".join(classes) + '"', allow_blank=False)
    dv.error = "Choose one of the four SCF classifications"
    dv.errorTitle = "Invalid classification"
    dv.prompt = "Select classification"
    dv.promptTitle = "SCF section"
    ws.add_data_validation(dv)
    dv.add("C8:C18")

    for i, (item, txn, clas, cash, why) in enumerate(rows):
        r = 8 + i
        ws.row_dimensions[r].height = 48
        fill = fill_white if i % 2 == 0 else fill_pale
        cA = ws.cell(r, 2, item)
        paint(cA, fill, Font(name="Calibri", size=12, bold=True, color=NAVY), center, thin)
        cB = ws.cell(r, 3, txn)
        paint(cB, fill, font_body, left, thin)
        cC = ws.cell(r, 4, clas)
        paint(cC, fill_green, font_ans, center, thin)
        cD = ws.cell(r, 5, cash)
        paint(cD, fill, font_body, center, thin)
        cE = ws.cell(r, 6, why)
        paint(cE, fill, Font(name="Calibri", size=10), left, thin)

    # legend
    header_bar(ws, 20, 2, 6, "Decision rules (memorise these)")
    rules = [
        ("Operating (indirect)", "Start with NI. Add non-cash expenses (depreciation, amortisation). Add losses / deduct gains on asset sales. Adjust working-capital current assets & current liabilities."),
        ("Investing", "Cash paid or received for long-lived assets and investments: buy/sell PPE, land, patents, shares of other companies, lending money / collecting loans."),
        ("Financing", "Cash with owners and long-term creditors: issue/retire shares, issue/retire bonds or notes, pay cash dividends."),
        ("Significant non-cash", "Material investing/financing that never touches cash: asset-for-asset swap, bonds issued for land, conversion of bonds to shares. Reported in a separate schedule, not in the three cash sections."),
    ]
    rr = 21
    for title, text in rules:
        ws.cell(rr, 2, title).font = font_label
        ws.cell(rr, 2).fill = fill_lt
        ws.cell(rr, 2).alignment = left
        ws.cell(rr, 2).border = thin
        ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=6)
        ws.cell(rr, 3, text).font = font_body
        ws.cell(rr, 3).alignment = left
        for c in range(3, 7):
            ws.cell(rr, c).border = thin
            ws.cell(rr, c).fill = fill_grey
        ws.row_dimensions[rr].height = 40
        rr += 1

    ws.merge_cells("B26:F26")
    ws["B26"] = (
        "Quick trap: a loss on disposal is Operating (add-back), while the cash proceeds from the same disposal are Investing. "
        "Do not put the loss itself in Investing."
    )
    ws["B26"].font = Font(name="Calibri", size=10, italic=True, color=RED)
    ws.row_dimensions[26].height = 28
    ws.print_title_rows = "1:7"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # ---- E17.12 ----
    ws = wb.create_sheet("E17.12 Direct Method")
    set_col_widths(ws, {c: w for c, w in zip("ABCDEFGHIJK", [3, 42, 16, 16, 3, 42, 16, 16, 3, 22, 18])})
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B2:H2")
    ws["B2"] = "E17.12 (LO 4)  ·  Compute cash payments — direct method"
    ws["B2"].font = font_title
    ws.merge_cells("B3:H3")
    ws["B3"] = "McDonald's Corporation  ·  Year 2027  ·  amounts in $ millions"
    ws["B3"].font = font_subtitle

    # given data
    header_bar(ws, 5, 2, 4, "Given data (from the tutorial PDF)")
    given = [
        ("Income statement — Cost of goods sold", 5178.0),
        ("Income statement — Operating expenses (incl. depreciation)", 10725.7),
        ("Included depreciation expense", 1216.2),
        ("Inventory decreased", 5.3),
        ("Prepaid expenses increased", 42.2),
        ("Accounts payable (inventory suppliers) increased", 15.6),
        ("Accrued expenses payable increased", 199.8),
    ]
    ws["B6"] = "Item"
    ws["C6"] = "$ millions"
    paint(ws["B6"], fill_blue, font_h, center, thin)
    paint(ws["C6"], fill_blue, font_h, center, thin)
    labels_cells = []
    for i, (lab, val) in enumerate(given):
        r = 7 + i
        ws.cell(r, 2, lab).font = font_body
        ws.cell(r, 2).fill = fill_lt if i % 2 == 0 else fill_white
        ws.cell(r, 2).border = thin
        cell = ws.cell(r, 3, val)
        money(cell)
        cell.fill = fill_yellow
        cell.border = thin
        cell.alignment = Alignment(horizontal="right", vertical="center")

    # named-range-like absolute refs: C7..C13
    # C7 COGS, C8 OpEx, C9 Dep, C10 Inv decr, C11 Prepaid incr, C12 AP incr, C13 Accr incr

    # Part (a)
    header_bar(ws, 16, 2, 4, "(a)  Cash payments to suppliers")
    ws["B17"] = "Line"
    ws["C17"] = "Amount"
    ws["D17"] = "Direction"
    for col in range(2, 5):
        paint(ws.cell(17, col), fill_blue, font_h, center, thin)

    a_lines = [
        ("Cost of goods sold", "=C7", "start"),
        ("Less: Decrease in inventory", "=C10", "inventory ↓ means purchases < COGS"),
        ("Purchases", "=C18-C19", "goods bought this year"),
        ("Less: Increase in accounts payable", "=C12", "AP ↑ means we have not yet paid all purchases"),
        ("Cash payments to suppliers", "=C20-C21", "ANSWER (a)"),
    ]
    for i, (lab, formula, note) in enumerate(a_lines):
        r = 18 + i
        ws.cell(r, 2, lab).font = font_label if i in (2, 4) else font_body
        ws.cell(r, 2).fill = fill_green if i == 4 else (fill_pale if i == 2 else fill_white)
        ws.cell(r, 2).border = thin
        cell = ws.cell(r, 3, formula)
        money(cell)
        cell.border = bottom_double if i == 4 else (bottom_single if i == 2 else thin)
        cell.fill = fill_green if i == 4 else fill_white
        if i == 4:
            cell.font = Font(name="Calibri", size=12, bold=True, color=DKGREEN)
        ws.cell(r, 4, note).font = font_small
        ws.cell(r, 4).alignment = left
        ws.cell(r, 4).border = thin

    ws.merge_cells("B24:D24")
    ws["B24"] = "Formula:  Cash paid to suppliers  =  COGS  −  Decrease in inventory  −  Increase in AP"
    ws["B24"].font = font_formula
    ws.merge_cells("B25:D25")
    ws["B25"] = "Check:  5,178.0 − 5.3 − 15.6  =  5,157.1     If inventory had increased we would ADD it; if AP had decreased we would ADD it."
    ws["B25"].font = font_small

    # Part (b)
    header_bar(ws, 27, 2, 4, "(b)  Cash payments for operating expenses")
    ws["B28"] = "Line"
    ws["C28"] = "Amount"
    ws["D28"] = "Direction"
    for col in range(2, 5):
        paint(ws.cell(28, col), fill_blue, font_h, center, thin)

    b_lines = [
        ("Operating expenses (as reported)", "=C8", "includes depreciation"),
        ("Less: Depreciation expense", "=C9", "non-cash — never paid in cash"),
        ("Cash-basis operating expenses before WC", "=C29-C30", "accrual opex stripped of depreciation"),
        ("Add: Increase in prepaid expenses", "=C11", "prepaid ↑ means extra cash paid in advance"),
        ("Subtotal", "=C31+C32", ""),
        ("Less: Increase in accrued expenses payable", "=C13", "accrued ↑ means some expenses not yet paid"),
        ("Cash payments for operating expenses", "=C33-C34", "ANSWER (b)"),
    ]
    for i, (lab, formula, note) in enumerate(b_lines):
        r = 29 + i
        is_ans = i == 6
        is_mid = i in (2, 4)
        ws.cell(r, 2, lab).font = font_label if is_ans or is_mid else font_body
        ws.cell(r, 2).fill = fill_green if is_ans else (fill_pale if is_mid else fill_white)
        ws.cell(r, 2).border = thin
        cell = ws.cell(r, 3, formula)
        money(cell)
        cell.border = bottom_double if is_ans else (bottom_single if is_mid else thin)
        cell.fill = fill_green if is_ans else fill_white
        if is_ans:
            cell.font = Font(name="Calibri", size=12, bold=True, color=DKGREEN)
        ws.cell(r, 4, note).font = font_small
        ws.cell(r, 4).alignment = left
        ws.cell(r, 4).border = thin

    ws.merge_cells("B37:D37")
    ws["B37"] = "Formula:  Cash paid for opex  =  Operating expenses  −  Depreciation  +  Increase in prepaids  −  Increase in accrued expenses"
    ws["B37"].font = font_formula
    ws.merge_cells("B38:D38")
    ws["B38"] = "Check:  10,725.7 − 1,216.2 + 42.2 − 199.8  =  9,351.9"
    ws["B38"].font = font_small

    # Answers box
    header_bar(ws, 40, 2, 4, "Final answers")
    ws["B41"] = "(a) Cash payments to suppliers"
    ws["C41"] = "=C22"
    money(ws["C41"])
    ws["D41"] = "million"
    ws["B42"] = "(b) Cash payments for operating expenses"
    ws["C42"] = "=C35"
    money(ws["C42"])
    ws["D42"] = "million"
    for r in (41, 42):
        ws.cell(r, 2).font = font_label
        ws.cell(r, 2).fill = fill_lt
        ws.cell(r, 2).border = thin
        ws.cell(r, 3).fill = fill_green
        ws.cell(r, 3).font = Font(name="Calibri", size=14, bold=True, color=DKGREEN)
        ws.cell(r, 3).border = med
        ws.cell(r, 4).border = thin
        ws.row_dimensions[r].height = 22

    # T-account illustration on the right
    header_bar(ws, 5, 6, 8, "T-account reconstruction (same numbers)")
    ws.merge_cells("F6:H6")
    ws["F6"] = "Inventory"
    ws["F6"].font = font_h
    ws["F6"].fill = fill_blue
    ws["F6"].alignment = center
    ws["G6"].fill = fill_blue
    ws["H6"].fill = fill_blue
    ws["F7"] = "Purchases (plug)"
    ws["G7"] = "=C20"
    money(ws["G7"])
    ws["H7"] = "COGS"
    ws["F8"] = ""
    ws["G8"] = ""
    ws["H8"] = "=C7"
    money(ws["H8"])
    ws["F9"] = "Decrease"
    ws["G9"] = "=C10"
    money(ws["G9"])
    for r in range(6, 10):
        for c in range(6, 9):
            if ws.cell(r, c).border.left.style is None:
                ws.cell(r, c).border = thin

    ws.merge_cells("F11:H11")
    ws["F11"] = "Accounts Payable (suppliers)"
    ws["F11"].font = font_h
    ws["F11"].fill = fill_blue
    ws["F11"].alignment = center
    ws["G11"].fill = fill_blue
    ws["H11"].fill = fill_blue
    ws["F12"] = "Cash paid (plug)"
    ws["G12"] = "=C22"
    money(ws["G12"])
    ws["H12"] = "Purchases"
    ws["F13"] = ""
    ws["H13"] = "=C20"
    money(ws["H13"])
    ws["F14"] = ""
    ws["H14"] = "AP increase"
    ws["G15"] = ""
    ws["H15"] = "=C12"
    money(ws["H15"])
    for r in range(11, 16):
        for c in range(6, 9):
            ws.cell(r, c).border = thin

    ws.merge_cells("F17:H20")
    ws["F17"] = (
        "Memory aid for current assets / liabilities\n"
        "Asset ↑  → subtract (more cash tied up)\n"
        "Asset ↓  → add (cash released)\n"
        "Liability ↑  → add to NI / subtract from cash paid (we deferred payment)\n"
        "Liability ↓  → subtract from NI / add to cash paid (we paid extra)"
    )
    ws["F17"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws["F17"].font = Font(name="Calibri", size=10)
    ws["F17"].fill = fill_orange
    for r in range(17, 21):
        for c in range(6, 9):
            ws.cell(r, c).border = thin
            ws.cell(r, c).fill = fill_orange

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_area = "A1:H43"

    wb.save("/workspace/ENGG433_Week04/Cashflow_statement_template.xlsx")
    print("wrote Cashflow_statement_template.xlsx")


# ---------------------------------------------------------------------------
# Workbook 2: E18.4 vertical analysis
# ---------------------------------------------------------------------------
def build_e184():
    wb = Workbook()
    ws = wb.active
    ws.title = "E18.4 Vertical Analysis"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 4, "B": 32, "C": 16, "D": 14, "E": 16, "F": 14, "G": 28, "H": 16})

    ws.merge_cells("B2:F2")
    ws["B2"] = "JOSHUA CORPORATION"
    ws["B2"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["B2"].alignment = center
    ws.merge_cells("B3:F3")
    ws["B3"] = "Condensed Income Statement"
    ws["B3"].font = Font(name="Calibri", size=13, bold=True, color=BLUE)
    ws["B3"].alignment = center
    ws.merge_cells("B4:F4")
    ws["B4"] = "For the Years Ended December 31"
    ws["B4"].font = font_subtitle
    ws["B4"].alignment = center
    ws.merge_cells("B5:F5")
    ws["B5"] = "E18.4 (LO 2)  ·  Prepare vertical analysis  ·  percentages of net sales, rounded to 1 decimal place"
    ws["B5"].font = font_small
    ws["B5"].alignment = center

    # year headers
    ws.merge_cells("C7:D7")
    ws["C7"] = "2027"
    ws.merge_cells("E7:F7")
    ws["E7"] = "2026"
    for cell in (ws["C7"], ws["E7"]):
        cell.font = font_h
        cell.fill = fill_navy
        cell.alignment = center
        cell.border = med
    ws["D7"].fill = fill_navy
    ws["D7"].border = med
    ws["F7"].fill = fill_navy
    ws["F7"].border = med
    ws["B7"].fill = fill_navy
    ws["B7"].border = med

    for col, txt in [(2, ""), (3, "Amount"), (4, "Percent"), (5, "Amount"), (6, "Percent")]:
        cell = ws.cell(8, col, txt)
        paint(cell, fill_blue, font_h, center, thin)

    # Input amounts (yellow) — only the given lines; derived lines are formulas
    # Row map:
    # 9 Net sales
    # 10 COGS
    # 11 Gross profit
    # 12 Selling
    # 13 Admin
    # 14 Total operating expenses
    # 15 Income before income taxes
    # 16 Income tax expense
    # 17 Net income

    def amt_fmt(cell):
        cell.number_format = '$#,##0;($#,##0);"—"'
        cell.alignment = Alignment(horizontal="right", vertical="center")

    # Net sales inputs
    ws["B9"] = "Net sales"
    ws["C9"] = 800000
    ws["E9"] = 600000
    ws["D9"] = "=IF($C$9=0,0,C9/$C$9)"
    ws["F9"] = "=IF($E$9=0,0,E9/$E$9)"

    ws["B10"] = "Cost of goods sold"
    ws["C10"] = 520000
    ws["E10"] = 408000
    ws["D10"] = "=C10/$C$9"
    ws["F10"] = "=E10/$E$9"

    ws["B11"] = "Gross profit"
    ws["C11"] = "=C9-C10"
    ws["E11"] = "=E9-E10"
    ws["D11"] = "=C11/$C$9"
    ws["F11"] = "=E11/$E$9"

    ws["B12"] = "Selling expenses"
    ws["C12"] = 120000
    ws["E12"] = 72000
    ws["D12"] = "=C12/$C$9"
    ws["F12"] = "=E12/$E$9"

    ws["B13"] = "Administrative expenses"
    ws["C13"] = 60000
    ws["E13"] = 48000
    ws["D13"] = "=C13/$C$9"
    ws["F13"] = "=E13/$E$9"

    ws["B14"] = "Total operating expenses"
    ws["C14"] = "=C12+C13"
    ws["E14"] = "=E12+E13"
    ws["D14"] = "=C14/$C$9"
    ws["F14"] = "=E14/$E$9"

    ws["B15"] = "Income before income taxes"
    ws["C15"] = "=C11-C14"
    ws["E15"] = "=E11-E14"
    ws["D15"] = "=C15/$C$9"
    ws["F15"] = "=E15/$E$9"

    ws["B16"] = "Income tax expense"
    ws["C16"] = 30000
    ws["E16"] = 24000
    ws["D16"] = "=C16/$C$9"
    ws["F16"] = "=E16/$E$9"

    ws["B17"] = "Net income"
    ws["C17"] = "=C15-C16"
    ws["E17"] = "=E15-E16"
    ws["D17"] = "=C17/$C$9"
    ws["F17"] = "=E17/$E$9"

    given_rows = {9, 10, 12, 13, 16}
    total_rows = {11, 14, 15}
    for r in range(9, 18):
        ws.row_dimensions[r].height = 20
        lab = ws.cell(r, 2)
        lab.font = Font(name="Calibri", size=11, bold=(r in (11, 14, 15, 17)))
        lab.alignment = left
        lab.border = thin
        if r == 17:
            lab.fill = fill_green
        elif r in total_rows:
            lab.fill = fill_pale
        else:
            lab.fill = fill_white
        for col in (3, 5):
            cell = ws.cell(r, col)
            amt_fmt(cell)
            cell.border = bottom_double if r == 17 else (bottom_single if r in (11, 14, 15) else thin)
            if r in given_rows:
                cell.fill = fill_yellow
            elif r == 17:
                cell.fill = fill_green
                cell.font = Font(name="Calibri", size=11, bold=True)
            else:
                cell.fill = fill_pale
                cell.font = Font(name="Calibri", size=11, bold=True)
        for col in (4, 6):
            cell = ws.cell(r, col)
            pct(cell)
            cell.border = bottom_double if r == 17 else (bottom_single if r in (11, 14, 15) else thin)
            cell.fill = fill_green if r == 17 else fill_lt
            if r in (11, 14, 15, 17):
                cell.font = Font(name="Calibri", size=11, bold=True)

    # indent expense lines
    for r in (12, 13, 16):
        ws.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells("B19:F19")
    ws["B19"] = (
        "Yellow cells = given amounts from the question. Green / blue percent columns = Amount ÷ Net sales. "
        "Excel stores 0.088 and displays 8.8% (30,000/800,000 = 3.75% → 3.8%; 70,000/800,000 = 8.75% → 8.8%)."
    )
    ws["B19"].font = font_small
    ws.row_dimensions[19].height = 32

    # Interpretation
    header_bar(ws, 21, 2, 6, "What the vertical analysis tells us (2026 → 2027)")
    insights = [
        ("Gross margin improved", "COGS fell from 68.0% to 65.0% of sales, so gross profit rose from 32.0% to 35.0%. The company kept more of each sales dollar after product cost."),
        ("Selling costs jumped", "Selling expenses rose from 12.0% to 15.0% of sales — a 3 percentage-point deterioration, enough to offset most of the gross-margin gain."),
        ("Admin slightly better", "Administrative expenses eased from 8.0% to 7.5% of sales (fixed-cost leverage as sales grew 33%)."),
        ("Tax rate on sales down", "Income tax expense 4.0% → 3.8% of sales."),
        ("Bottom line up", "Net income margin 8.0% → 8.8%. Profitability improved, but the extra selling spend is the item to watch."),
    ]
    rr = 22
    for title, text in insights:
        ws.cell(rr, 2, title).font = font_label
        ws.cell(rr, 2).fill = fill_lt
        ws.cell(rr, 2).alignment = left
        ws.cell(rr, 2).border = thin
        ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=6)
        ws.cell(rr, 3, text).font = font_body
        ws.cell(rr, 3).alignment = left
        for c in range(3, 7):
            ws.cell(rr, c).border = thin
        ws.row_dimensions[rr].height = 36
        rr += 1

    # Comparison of key % 
    header_bar(ws, 28, 2, 6, "Key common-size percentages (chart data)")
    ws["B29"] = "Item"
    ws["C29"] = "2027"
    ws["D29"] = "2026"
    for col in range(2, 5):
        paint(ws.cell(29, col), fill_blue, font_h, center, thin)
    chart_items = [
        ("Cost of goods sold", "=D10", "=F10"),
        ("Gross profit", "=D11", "=F11"),
        ("Selling expenses", "=D12", "=F12"),
        ("Administrative expenses", "=D13", "=F13"),
        ("Income tax expense", "=D16", "=F16"),
        ("Net income", "=D17", "=F17"),
    ]
    for i, (lab, a, b) in enumerate(chart_items):
        r = 30 + i
        ws.cell(r, 2, lab).font = font_body
        ws.cell(r, 2).border = thin
        ws.cell(r, 3, a)
        ws.cell(r, 4, b)
        pct(ws.cell(r, 3))
        pct(ws.cell(r, 4))
        ws.cell(r, 3).border = thin
        ws.cell(r, 4).border = thin

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Vertical analysis: % of net sales"
    chart.y_axis.title = "% of net sales"
    chart.y_axis.scaling.min = 0
    chart.y_axis.numFmt = "0%"
    data = Reference(ws, min_col=3, min_row=29, max_col=4, max_row=35)
    cats = Reference(ws, min_col=2, min_row=30, max_row=35)
    chart.add_data(data, from_rows=False, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.style = 10
    chart.y_axis.numFmt = "0%"
    chart.width = 18
    chart.height = 9
    ws.add_chart(chart, "B38")

    # Data tab with given only (template feel)
    ws2 = wb.create_sheet("Given data")
    ws2.sheet_view.showGridLines = False
    set_col_widths(ws2, {"A": 4, "B": 32, "C": 16, "D": 16})
    ws2.merge_cells("B2:D2")
    ws2["B2"] = "Operating data for Joshua Corporation"
    ws2["B2"].font = font_title
    ws2["B3"] = "Paste / check the question figures here. The Vertical Analysis sheet reads these yellow cells."
    ws2["B3"].font = font_small
    ws2["C5"] = 2027
    ws2["D5"] = 2026
    paint(ws2["C5"], fill_navy, font_h, center, thin)
    paint(ws2["D5"], fill_navy, font_h, center, thin)
    paint(ws2["B5"], fill_navy, font_h, center, thin)
    given = [
        ("Net sales", 800000, 600000, "'E18.4 Vertical Analysis'!C9", "'E18.4 Vertical Analysis'!E9"),
        ("Cost of goods sold", 520000, 408000, "'E18.4 Vertical Analysis'!C10", "'E18.4 Vertical Analysis'!E10"),
        ("Selling expenses", 120000, 72000, "'E18.4 Vertical Analysis'!C12", "'E18.4 Vertical Analysis'!E12"),
        ("Administrative expenses", 60000, 48000, "'E18.4 Vertical Analysis'!C13", "'E18.4 Vertical Analysis'!E13"),
        ("Income tax expense", 30000, 24000, "'E18.4 Vertical Analysis'!C16", "'E18.4 Vertical Analysis'!E16"),
        ("Net income (check)", 70000, 48000, None, None),
    ]
    # Keep given data independent; vertical sheet already has values. This sheet is a readable source table.
    ws2["B6"] = "Item"
    ws2["C6"] = "2027"
    ws2["D6"] = "2026"
    for col in range(2, 5):
        paint(ws2.cell(6, col), fill_blue, font_h, center, thin)
    for i, row in enumerate(given):
        r = 7 + i
        ws2.cell(r, 2, row[0]).font = font_body
        ws2.cell(r, 2).border = thin
        ws2.cell(r, 3, row[1])
        ws2.cell(r, 4, row[2])
        amt_fmt(ws2.cell(r, 3))
        amt_fmt(ws2.cell(r, 4))
        ws2.cell(r, 3).fill = fill_yellow
        ws2.cell(r, 4).fill = fill_yellow
        ws2.cell(r, 3).border = thin
        ws2.cell(r, 4).border = thin

    ws2["B14"] = "Check: 2027 NI = 800,000 − 520,000 − 120,000 − 60,000 − 30,000 = 70,000"
    ws2["B14"].font = font_formula
    ws2["B15"] = "Check: 2026 NI = 600,000 − 408,000 − 72,000 − 48,000 − 24,000 = 48,000"
    ws2["B15"].font = font_formula

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_area = "A1:F36"

    wb.save("/workspace/ENGG433_Week04/excel_template_18.4.xlsx")
    print("wrote excel_template_18.4.xlsx")


# ---------------------------------------------------------------------------
# Workbook 3: E18.9 ratios
# ---------------------------------------------------------------------------
def build_e189():
    wb = Workbook()
    ws = wb.active
    ws.title = "E18.9 Ratios"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 3, "B": 38, "C": 16, "D": 16, "E": 3, "F": 42, "G": 18, "H": 16, "I": 28})

    ws.merge_cells("B2:D2")
    ws["B2"] = "LENDELL COMPANY"
    ws["B2"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["B2"].alignment = center
    ws.merge_cells("B3:D3")
    ws["B3"] = "Comparative Balance Sheets  ·  December 31"
    ws["B3"].font = font_subtitle
    ws["B3"].alignment = center
    ws.merge_cells("B4:D4")
    ws["B4"] = "E18.9 (LO 3)  ·  Compute selected ratios at 31 December 2027"
    ws["B4"].font = font_small
    ws["B4"].alignment = center

    ws["C6"] = 2027
    ws["D6"] = 2026
    paint(ws["B6"], fill_navy, font_h, center, thin)
    paint(ws["C6"], fill_navy, font_h, center, thin)
    paint(ws["D6"], fill_navy, font_h, center, thin)

    def amt(cell, value=None):
        if value is not None:
            cell.value = value
        cell.number_format = '$#,##0;($#,##0);"—"'
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = thin

    items = [
        ("Cash", 15000, 30000, True),
        ("Accounts receivable (net)", 70000, 60000, True),
        ("Inventory", 60000, 50000, True),
        ("Plant assets (net)", 200000, 180000, False),
        ("Total assets", None, None, False),
        ("Accounts payable", 50000, 60000, True),
        ("Bonds payable (15%, due 2040)", 100000, 100000, False),
        ("Common stock, $10 par", 140000, 120000, False),
        ("Retained earnings", 55000, 40000, False),
        ("Total liabilities & equity", None, None, False),
    ]
    # rows 7-16
    ws["B7"] = "Cash"
    amt(ws["C7"], 15000)
    amt(ws["D7"], 30000)
    ws["B8"] = "Accounts receivable (net)"
    amt(ws["C8"], 70000)
    amt(ws["D8"], 60000)
    ws["B9"] = "Inventory"
    amt(ws["C9"], 60000)
    amt(ws["D9"], 50000)
    ws["B10"] = "Plant assets (net)"
    amt(ws["C10"], 200000)
    amt(ws["D10"], 180000)
    ws["B11"] = "Total assets"
    ws["C11"] = "=C7+C8+C9+C10"
    ws["D11"] = "=D7+D8+D9+D10"
    amt(ws["C11"])
    amt(ws["D11"])
    ws["B12"] = "Accounts payable (current)"
    amt(ws["C12"], 50000)
    amt(ws["D12"], 60000)
    ws["B13"] = "Bonds payable 15% (due 2040)"
    amt(ws["C13"], 100000)
    amt(ws["D13"], 100000)
    ws["B14"] = "Common stock, $10 par"
    amt(ws["C14"], 140000)
    amt(ws["D14"], 120000)
    ws["B15"] = "Retained earnings"
    amt(ws["C15"], 55000)
    amt(ws["D15"], 40000)
    ws["B16"] = "Total liabilities & equity"
    ws["C16"] = "=C12+C13+C14+C15"
    ws["D16"] = "=D12+D13+D14+D15"
    amt(ws["C16"])
    amt(ws["D16"])

    for r in range(7, 17):
        ws.cell(r, 2).font = font_label if r in (11, 16) else font_body
        ws.cell(r, 2).border = thin
        ws.cell(r, 2).fill = fill_pale if r in (11, 16) else fill_white
        if r in (7, 8, 9, 10, 12, 13, 14, 15):
            ws.cell(r, 3).fill = fill_yellow
            ws.cell(r, 4).fill = fill_yellow
        else:
            ws.cell(r, 3).fill = fill_pale
            ws.cell(r, 4).fill = fill_pale
            ws.cell(r, 3).font = Font(name="Calibri", size=11, bold=True)
            ws.cell(r, 4).font = Font(name="Calibri", size=11, bold=True)
            ws.cell(r, 3).border = bottom_double
            ws.cell(r, 4).border = bottom_double

    header_bar(ws, 18, 2, 4, "Additional information for 2027")
    addl = [
        ("Net income", 25000),
        ("Sales on account", 375000),
        ("Sales returns and allowances", 25000),
        ("Cost of goods sold", 198000),
        ("Net cash provided by operating activities", 48000),
        ("Capital expenditures", 25000),
        ("Cash dividends paid", 10000),
    ]
    ws["B19"] = "Item"
    ws["C19"] = "Amount"
    paint(ws["B19"], fill_blue, font_h, center, thin)
    paint(ws["C19"], fill_blue, font_h, center, thin)
    for i, (lab, val) in enumerate(addl):
        r = 20 + i
        ws.cell(r, 2, lab).font = font_body
        ws.cell(r, 2).border = thin
        ws.cell(r, 3, val)
        amt(ws.cell(r, 3))
        ws.cell(r, 3).fill = fill_yellow

    ws["B28"] = "Net credit sales (Sales − returns)"
    ws["C28"] = "=C21-C22"
    amt(ws["C28"])
    ws["B28"].font = font_label
    ws["B28"].fill = fill_lt
    ws["C28"].fill = fill_pale
    ws["B29"] = "Average accounts receivable"
    ws["C29"] = "=(C8+D8)/2"
    amt(ws["C29"])
    ws["B29"].font = font_label
    ws["B29"].fill = fill_lt
    ws["C29"].fill = fill_pale
    ws["B30"] = "Average inventory"
    ws["C30"] = "=(C9+D9)/2"
    amt(ws["C30"])
    ws["B30"].font = font_label
    ws["B30"].fill = fill_lt
    ws["C30"].fill = fill_pale
    ws["B31"] = "Current assets (31 Dec 2027)"
    ws["C31"] = "=C7+C8+C9"
    amt(ws["C31"])
    ws["B31"].font = font_label
    ws["B31"].fill = fill_lt
    ws["C31"].fill = fill_pale
    ws["B32"] = "Current liabilities (AP only; bonds due 2040)"
    ws["C32"] = "=C12"
    amt(ws["C32"])
    ws["B32"].font = font_label
    ws["B32"].fill = fill_lt
    ws["C32"].fill = fill_pale

    # Ratio computations
    header_bar(ws, 6, 6, 9, "Required ratios at 31 December 2027")
    ws["F7"] = "Ratio"
    ws["G7"] = "Formula"
    ws["H7"] = "Answer"
    ws["I7"] = "Workings"
    for col in range(6, 10):
        paint(ws.cell(7, col), fill_blue, font_h, center, thin)

    ratios = [
        (
            "a. Current ratio",
            "Current assets ÷ Current liabilities",
            "=C31/C32",
            '0.0" : 1"',
            '="$"&TEXT(C31,"#,##0")&" / $"&TEXT(C32,"#,##0")&" = "&TEXT(H8,"0.0")&" : 1"',
        ),
        (
            "b. Accounts receivable turnover",
            "Net credit sales ÷ Average AR",
            "=C28/C29",
            '0.0" times"',
            '="$"&TEXT(C28,"#,##0")&" / $"&TEXT(C29,"#,##0")&" = "&TEXT(H9,"0.0")&" times"',
        ),
        (
            "c. Average collection period",
            "365 ÷ AR turnover (using 5.4)",
            "=365/ROUND(H9,1)",
            '0.0" days"',
            '="365 / "&TEXT(ROUND(H9,1),"0.0")&" = "&TEXT(H10,"0.0")&" days"',
        ),
        (
            "d. Inventory turnover",
            "COGS ÷ Average inventory",
            "=C23/C30",
            '0.0" times"',
            '="$"&TEXT(C23,"#,##0")&" / $"&TEXT(C30,"#,##0")&" = "&TEXT(H11,"0.0")&" times"',
        ),
        (
            "e. Days in inventory",
            "365 ÷ Inventory turnover (using 3.6)",
            "=365/ROUND(H11,1)",
            '0.0" days"',
            '="365 / "&TEXT(ROUND(H11,1),"0.0")&" = "&TEXT(H12,"0.0")&" days"',
        ),
        (
            "f. Free cash flow",
            "CFO − Capex − Cash dividends",
            "=C24-C25-C26",
            '"$"#,##0',
            '="$"&TEXT(C24,"#,##0")&" − $"&TEXT(C25,"#,##0")&" − $"&TEXT(C26,"#,##0")&" = $"&TEXT(H13,"#,##0")',
        ),
    ]
    for i, (name, formula, excel, numfmt, work) in enumerate(ratios):
        r = 8 + i
        ws.row_dimensions[r].height = 28
        ws.cell(r, 6, name).font = font_label
        ws.cell(r, 6).fill = fill_lt
        ws.cell(r, 6).border = thin
        ws.cell(r, 6).alignment = left
        ws.cell(r, 7, formula).font = Font(name="Calibri", size=10, italic=True)
        ws.cell(r, 7).border = thin
        ws.cell(r, 7).alignment = left
        cell = ws.cell(r, 8, excel)
        cell.number_format = numfmt
        cell.font = Font(name="Calibri", size=12, bold=True, color=DKGREEN)
        cell.fill = fill_green
        cell.alignment = center
        cell.border = med
        ws.cell(r, 9, work).font = Font(name="Calibri", size=10)
        ws.cell(r, 9).border = thin
        ws.cell(r, 9).alignment = left

    ws.merge_cells("F15:I18")
    ws["F15"] = (
        "Notes for the exam / tutorial:\n"
        "• Bonds payable are due in 2040, so they are NON-CURRENT. Current liabilities = Accounts payable only.\n"
        "• Net credit sales = Sales on account − sales returns and allowances = $375,000 − $25,000 = $350,000.\n"
        "• Turnover ratios use AVERAGES of beginning and ending balances.\n"
        "• Collection period and days in inventory use the turnover already rounded to 1 decimal (Wiley convention): 365/5.4 = 67.6 days; 365/3.6 = 101.4 days.\n"
        "• Free cash flow is not a ratio; it is a dollar amount of cash left after maintaining operating capacity and paying dividends."
    )
    ws["F15"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["F15"].font = Font(name="Calibri", size=10)
    ws["F15"].fill = fill_orange
    for r in range(15, 19):
        for c in range(6, 10):
            ws.cell(r, c).border = thin
            ws.cell(r, c).fill = fill_orange

    header_bar(ws, 20, 6, 9, "Interpretation (one line each)")
    interps = [
        "Current ratio 2.9 : 1 — $2.90 of current assets for every $1 of current liabilities; short-term liquidity looks comfortable.",
        "AR turns 5.4 times — receivables are collected about 5.4 times a year (~every 68 days). Compare with credit terms.",
        "Inventory turns 3.6 times — stock sits about 101 days. Slow if the firm is a retailer of perishable/fast goods; maybe normal for specialised equipment.",
        "FCF $13,000 — operations generated $48,000; after $25,000 capex and $10,000 dividends, $13,000 remains to repay debt or hold as cash.",
    ]
    for i, t in enumerate(interps):
        r = 21 + i
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=9)
        ws.cell(r, 6, t).font = font_body
        ws.cell(r, 6).alignment = left
        for c in range(6, 10):
            ws.cell(r, c).border = thin
        ws.row_dimensions[r].height = 32

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    wb.save("/workspace/ENGG433_Week04/E18.9_Lendell_ratios.xlsx")
    print("wrote E18.9_Lendell_ratios.xlsx")


if __name__ == "__main__":
    build_cashflow()
    build_e184()
    build_e189()
