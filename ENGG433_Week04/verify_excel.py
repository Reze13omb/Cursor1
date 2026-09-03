#!/usr/bin/env python3
"""Evaluate Week 4 Excel formulas without Excel and assert expected answers."""

from __future__ import annotations

import math
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path("/workspace/ENGG433_Week04")
CELL = re.compile(r"\$?([A-Z]+)\$?(\d+)")


def col_row(addr: str) -> tuple[str, int]:
    m = CELL.fullmatch(addr.replace("$", ""))
    if not m:
        raise ValueError(addr)
    return m.group(1), int(m.group(2))


class SheetCalc:
    def __init__(self, ws):
        self.raw = {}
        self.cache = {}
        self.visiting = set()
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    self.raw[cell.coordinate] = cell.value

    def get(self, addr: str):
        addr = addr.replace("$", "")
        if addr in self.cache:
            return self.cache[addr]
        if addr in self.visiting:
            raise RuntimeError(f"cycle at {addr}")
        if addr not in self.raw:
            return 0
        val = self.raw[addr]
        if not (isinstance(val, str) and val.startswith("=")):
            self.cache[addr] = val
            return val
        self.visiting.add(addr)
        out = self.eval_formula(val[1:])
        self.visiting.remove(addr)
        self.cache[addr] = out
        return out

    def eval_formula(self, expr: str):
        expr = expr.strip()
        # IF(cond, a, b)
        if expr.upper().startswith("IF(") and expr.endswith(")"):
            inner = expr[3:-1]
            parts = self._split_args(inner)
            cond = parts[0]
            if "=" in cond:
                left, right = cond.split("=", 1)
                ok = self.eval_formula(left) == self.eval_formula(right)
            else:
                ok = bool(self.eval_formula(cond))
            return self.eval_formula(parts[1] if ok else parts[2])
        if expr.upper().startswith("ROUND("):
            inner = expr[expr.find("(") + 1 : -1]
            a, b = self._split_args(inner)
            return round(float(self.eval_formula(a)), int(self.eval_formula(b)))
        # replace cell refs with values — longest first
        refs = sorted({m.group(0) for m in CELL.finditer(expr)}, key=len, reverse=True)
        py = expr
        for ref in refs:
            v = self.get(ref.replace("$", ""))
            py = py.replace(ref, f"({v})")
        py = py.replace("/0", "/0.0")
        return eval(
            py,
            {"__builtins__": {}},
            {"ROUND": lambda n, d: round(float(n), int(d))},
        )

    @staticmethod
    def _split_args(s: str):
        parts, buf, depth = [], [], 0
        for ch in s:
            if ch == "," and depth == 0:
                parts.append("".join(buf))
                buf = []
                continue
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
            buf.append(ch)
        parts.append("".join(buf))
        return parts


def near(a, b, eps=1e-6):
    return abs(float(a) - float(b)) <= eps


def main():
    failures = []
    lines = []

    def check(label, got, exp):
        ok = near(got, exp)
        status = "PASS" if ok else "FAIL"
        lines.append(f"{status:4}  {label:<42} got={got}  expected={exp}")
        if not ok:
            failures.append(label)

    # --- E17.12 ---
    wb = load_workbook(ROOT / "Cashflow_statement_template.xlsx", data_only=False)
    calc = SheetCalc(wb["E17.12 Direct Method"])
    check("E17.12 COGS input C7", calc.get("C7"), 5178.0)
    check("E17.12 purchases C20", calc.get("C20"), 5172.7)
    check("E17.12 cash to suppliers C22", calc.get("C22"), 5157.1)
    check("E17.12 cash opex C35", calc.get("C35"), 9351.9)
    check("E17.12 answer box C41", calc.get("C41"), 5157.1)
    check("E17.12 answer box C42", calc.get("C42"), 9351.9)
    check("E17.12 T-account G12", calc.get("G12"), 5157.1)

    # E17.2 classifications
    ws = wb["E17.2 Classification"]
    expected_cls = {
        8: "Significant noncash investing and financing activity",
        9: "Investing activity",
        10: "Financing activity",
        11: "Operating activity",
        12: "Significant noncash investing and financing activity",
        13: "Financing activity",
        14: "Operating activity",
        15: "Significant noncash investing and financing activity",
        16: "Investing activity",
        17: "Operating activity",
        18: "Financing activity",
    }
    for row, exp in expected_cls.items():
        got = ws.cell(row, 4).value
        ok = got == exp
        lines.append(f"{'PASS' if ok else 'FAIL':4}  E17.2 row {row}  {got}")
        if not ok:
            failures.append(f"E17.2 row {row}")

    # --- E18.4 ---
    wb = load_workbook(ROOT / "excel_template_18.4.xlsx", data_only=False)
    calc = SheetCalc(wb["E18.4 Vertical Analysis"])
    check("E18.4 GP 2027 C11", calc.get("C11"), 280000)
    check("E18.4 opex 2027 C14", calc.get("C14"), 180000)
    check("E18.4 IBT 2027 C15", calc.get("C15"), 100000)
    check("E18.4 NI 2027 C17", calc.get("C17"), 70000)
    check("E18.4 GP 2026 E11", calc.get("E11"), 192000)
    check("E18.4 NI 2026 E17", calc.get("E17"), 48000)
    check("E18.4 % COGS 2027 D10", round(calc.get("D10") * 100, 1), 65.0)
    check("E18.4 % GP 2027 D11", round(calc.get("D11") * 100, 1), 35.0)
    check("E18.4 % sell 2027 D12", round(calc.get("D12") * 100, 1), 15.0)
    check("E18.4 % admin 2027 D13", round(calc.get("D13") * 100, 1), 7.5)
    check("E18.4 % tax 2027 D16", round(calc.get("D16") * 100, 1), 3.8)
    check("E18.4 % NI 2027 D17", round(calc.get("D17") * 100, 1), 8.8)
    check("E18.4 % COGS 2026 F10", round(calc.get("F10") * 100, 1), 68.0)
    check("E18.4 % NI 2026 F17", round(calc.get("F17") * 100, 1), 8.0)

    # --- E18.9 ---
    wb = load_workbook(ROOT / "E18.9_Lendell_ratios.xlsx", data_only=False)
    calc = SheetCalc(wb["E18.9 Ratios"])
    check("E18.9 total assets 2027 C11", calc.get("C11"), 345000)
    check("E18.9 total L+E 2027 C16", calc.get("C16"), 345000)
    check("E18.9 total assets 2026 D11", calc.get("D11"), 320000)
    check("E18.9 total L+E 2026 D16", calc.get("D16"), 320000)
    check("E18.9 net credit sales C28", calc.get("C28"), 350000)
    check("E18.9 avg AR C29", calc.get("C29"), 65000)
    check("E18.9 avg inventory C30", calc.get("C30"), 55000)
    check("E18.9 current assets C31", calc.get("C31"), 145000)
    check("E18.9 current liabilities C32", calc.get("C32"), 50000)
    check("E18.9 current ratio H8", round(calc.get("H8"), 1), 2.9)
    check("E18.9 AR turnover H9", round(calc.get("H9"), 1), 5.4)
    check("E18.9 collection period H10", round(calc.get("H10"), 1), 67.6)
    check("E18.9 inventory turnover H11", round(calc.get("H11"), 1), 3.6)
    check("E18.9 days in inventory H12", round(calc.get("H12"), 1), 101.4)
    check("E18.9 free cash flow H13", calc.get("H13"), 13000)

    text = "\n".join(lines)
    print(text)
    print()
    if failures:
        print("FAILED:", ", ".join(failures))
        raise SystemExit(1)
    print(f"ALL {len(lines)} CHECKS PASSED")


if __name__ == "__main__":
    main()
